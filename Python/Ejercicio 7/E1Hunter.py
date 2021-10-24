# KEILA SOFIA CABALLERO RAMOS
# JOSE LUIS HERNANDEZ MEZA
# IRVIN MARTINEZ GONZALEZ
# FRANCISCO JAVIER VALERIO LARA

# IMPORTACIONES
from pyhunter import PyHunter
from openpyxl import Workbook
import getpass


# BUSQUEDA EN API
def busqueda(organizacion):
    resultado = hunter.domain_search(company=organizacion, limit=1,
                                     emails_type='personal')
    return resultado


# GUARDAR INFO EN EXCEL
def guardar_informacion(datos_encontrados, organizacion):
    libro = Workbook()
    hoja = libro.create_sheet(organizacion)
    libro.save("Hunter" + organizacion + ".xlsx")
    hoja.cell(1, 1, "Dominio")
    hoja.cell(1, 2, datos_encontrados['domain'])
    hoja.cell(2, 1, "Organización")
    hoja.cell(2, 2, datos_encontrados['organization'])
    hoja.cell(3, 1, "Correo")
    hoja.cell(3, 2, datos_encontrados['emails'][0]['value'])
    hoja.cell(4, 1, "Nombre(s)")
    hoja.cell(4, 2, datos_encontrados['emails'][0]['first_name'])
    hoja.cell(5, 1, "Apellidos")
    hoja.cell(5, 2, datos_encontrados['emails'][0]['last_name'])
    hoja.cell(6, 1, "Recursos")
    hoja.cell(6, 2, datos_encontrados['emails'][0]['sources'][0]['domain'])
    hoja.cell(6, 3, datos_encontrados['emails'][0]['sources'][0]['uri'])
    libro.save("Hunter" + organizacion + ".xlsx")


# RESULTADO EN TERMINAL
print("Script para buscar información")
apikey = getpass.getpass("Ingresa tu API key: ")
hunter = PyHunter(apikey)
orga = input("Dominio a investigar: ")
datos_encontrados = busqueda(orga)
if datos_encontrados == 'if cond is None:':
    exit()
else:
    print(datos_encontrados)
    print(type(datos_encontrados))
    guardar_informacion(datos_encontrados, orga)
